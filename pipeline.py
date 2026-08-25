import os
import re
import io
import json
import unicodedata
from statistics import median

import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Konfigurace
# ---------------------------------------------------------------------------

STORE_DEFAULT_VENDOR = os.environ.get("STORE_DEFAULT_VENDOR", "Northfield")
STORE_NAME_FOR_SEO = os.environ.get("STORE_NAME_FOR_SEO", STORE_DEFAULT_VENDOR)

ANTHROPIC_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-5")
AI_BATCH_SIZE = int(os.environ.get("AI_BATCH_SIZE", "5"))

COLUMN_ALIASES = {
    "product name": "title",
    "title": "title",
    "name": "title",
    "brand": "vendor",
    "vendor": "vendor",
    "category": "category",
    "short description": "description",
    "description": "description",
    "price": "price",
    "currency": "currency",
    "sku": "sku",
    "stock": "inventory_qty",
    "stock quantity": "inventory_qty",
    "weight_g": "weight_g",
    "weight (g)": "weight_g",
    "image url": "image_url",
    "image": "image_url",
    "color": "color",
    "material": "material",
}

# Zaloha pro sloupce, ktere neodpovidaji presnemu aliasu (napr. vlastni export
# s pripojenym suffixem jako "Price (input)" nebo jinak pojmenovany vstup) -
# hledani podle klicoveho slova v nazvu sloupce, ne podle presneho retezce.
FIELD_KEYWORDS = [
    ("title", ["product name", "title", "name"]),
    ("vendor", ["vendor", "brand", "manufacturer", "znacka", "vyrobce"]),
    ("category", ["category", "kategorie"]),
    ("description", ["description", "popis"]),
    ("price", ["price", "cena"]),
    ("sku", ["sku"]),
    ("inventory_qty", ["stock", "inventory", "quantity", "qty", "sklad"]),
    ("weight_g", ["weight", "hmotnost", "vaha"]),
    ("image_url", ["image", "img", "photo", "picture", "obrazek"]),
    ("color", ["color", "colour", "barva"]),
    ("material", ["material"]),
]


def _normalize_column_key(col) -> str:
    key = str(col).strip().lower()
    key = re.sub(r"\([^)]*\)", " ", key)  # strip parenthetical qualifiers, e.g. "(input)"
    key = re.sub(r"[_\-]+", " ", key)
    return re.sub(r"\s+", " ", key).strip()

CATEGORY_KEYWORDS = [
    # obleceni (puvodni demo katalog)
    ("shirt", "Shirts"), ("tee", "T-Shirts"), ("sweater", "Sweaters"),
    ("pullover", "Sweaters"), ("sweatshirt", "Sweaters"), ("jacket", "Jackets"),
    ("windbreaker", "Jackets"), ("trouser", "Trousers"), ("pant", "Trousers"),
    ("short", "Activewear"), ("bottle", "Accessories"), ("bag", "Accessories"),
    ("tote", "Accessories"), ("wallet", "Accessories"), ("card holder", "Accessories"),
    ("cap", "Hats"), ("beanie", "Hats"), ("hat", "Hats"),
    # domacnost / kucyhnske potreby
    ("pánev", "Nádobí"), ("panev", "Nádobí"), ("wok", "Nádobí"), ("tegame", "Nádobí"),
    ("hrnec", "Nádobí"), ("kastrol", "Nádobí"), ("rendlík", "Nádobí"), ("rendlik", "Nádobí"),
    ("pekáč", "Nádobí"), ("pekac", "Nádobí"), ("zapékací", "Nádobí"), ("zapekaci", "Nádobí"),
    ("nůž", "Nože a příbory"), ("nuz", "Nože a příbory"), ("ocílka", "Nože a příbory"),
    ("ocilka", "Nože a příbory"), ("příbor", "Nože a příbory"), ("pribor", "Nože a příbory"),
    ("vidlič", "Nože a příbory"), ("vidlic", "Nože a příbory"),
    ("prkénko", "Kuchyňské náčiní"), ("prkenko", "Kuchyňské náčiní"), ("kleště", "Kuchyňské náčiní"),
    ("kleste", "Kuchyňské náčiní"), ("hmoždíř", "Kuchyňské náčiní"), ("hmozdir", "Kuchyňské náčiní"),
    ("mísa", "Kuchyňské náčiní"), ("misa", "Kuchyňské náčiní"), ("mlékovar", "Kuchyňské náčiní"),
    ("mlekovar", "Kuchyňské náčiní"), ("napařovací", "Kuchyňské náčiní"), ("naparovaci", "Kuchyňské náčiní"),
    ("poukaz", "Dárkové poukazy"), ("voucher", "Dárkové poukazy"),
]


def get_ai_mode() -> str:
    return "anthropic" if os.environ.get("ANTHROPIC_API_KEY") else "mock"


# ---------------------------------------------------------------------------
# Nacteni a normalizace vstupu
# ---------------------------------------------------------------------------

def load_input(path_or_buffer) -> pd.DataFrame:
    df = pd.read_excel(path_or_buffer)
    rename_map = {}
    claimed = set()

    # 1. presny alias po odstraneni napr. "(input)"/"(g)" apod.
    for col in df.columns:
        key = _normalize_column_key(col)
        target = COLUMN_ALIASES.get(key)
        if target and target not in claimed:
            rename_map[col] = target
            claimed.add(target)

    # 2. zbytek: hledani klicoveho slova v nazvu sloupce (prvni shoda vyhrava,
    # takze uz obsazene pole se znovu neprepise - napr. "AI Suggested Category"
    # nesebere misto sloupci "Category")
    for col in df.columns:
        if col in rename_map:
            continue
        key = _normalize_column_key(col)
        for target, keywords in FIELD_KEYWORDS:
            if target in claimed:
                continue
            if any(kw in key for kw in keywords):
                rename_map[col] = target
                claimed.add(target)
                break

    df = df.rename(columns=rename_map)

    required_internal = ["title", "vendor", "category", "description", "price",
                          "sku", "inventory_qty", "weight_g", "image_url", "color", "material"]
    for col in required_internal:
        if col not in df.columns:
            df[col] = None
    return df


def clean_price(raw):
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw)
    match = re.search(r"-?\d+([.,]\d+)?", text.replace(" ", ""))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", "."))
    except ValueError:
        return None


def clean_text(val) -> str:
    """Bezpecne prevede hodnotu na trimovany retezec; osetruje NaN z
    prazdnych bunek Excelu (ktere je v Pythonu 'truthy')."""
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return re.sub(r"-+", "-", text).strip("-")


# ---------------------------------------------------------------------------
# Validace
# ---------------------------------------------------------------------------

def _auto_dedupe_skus(df: pd.DataFrame) -> dict:
    """Automaticky odlisi duplicitni SKU pripojenim pripony (-2, -3, ...) misto
    aby duplicita blokovala import - SKU se jen upravi, nevymysli se od znova.
    Mutuje `df` na miste. Vraci {index: puvodni_sku} pro radky, ktere byly
    upraveny (pro nasledne oznaceni jako varovani k potvrzeni)."""
    sku_series = df["sku"].apply(clean_text)
    seen_counts = {}
    used_skus = set(s for s in sku_series if s)
    original_by_idx = {}

    for idx in df.index:
        sku = sku_series.loc[idx]
        if not sku:
            continue
        seen_counts[sku] = seen_counts.get(sku, 0) + 1
        if seen_counts[sku] == 1:
            continue  # prvni vyskyt zustava beze zmeny
        suffix = seen_counts[sku]
        new_sku = f"{sku}-{suffix}"
        while new_sku in used_skus:
            suffix += 1
            new_sku = f"{sku}-{suffix}"
        df.at[idx, "sku"] = new_sku
        used_skus.add(new_sku)
        original_by_idx[idx] = sku

    return original_by_idx


def validate_rows(df: pd.DataFrame):
    results = []
    auto_deduped = _auto_dedupe_skus(df)

    sku_series = df["sku"].apply(clean_text)
    sku_counts = sku_series[sku_series != ""].value_counts()

    handle_counts = {}
    for _, row in df.iterrows():
        h = slugify(clean_text(row.get("title")))
        if h:
            handle_counts[h] = handle_counts.get(h, 0) + 1

    for idx, row in df.iterrows():
        hard, soft = [], []
        # ktera pole (sloupce) presne zpusobila problem - pro cilene zvyrazneni
        # bunky misto celeho radku. Drzeny oddelene, protoze blokujici a
        # varovaci problem u stejneho pole se maji zvyraznit jinou barvou.
        hard_fields, soft_fields = set(), set()

        title = clean_text(row.get("title"))
        vendor = clean_text(row.get("vendor"))
        category = clean_text(row.get("category"))
        description = clean_text(row.get("description"))
        sku = clean_text(row.get("sku"))
        image_url = clean_text(row.get("image_url"))
        price_val = clean_price(row.get("price"))
        inv_qty = row.get("inventory_qty")

        if not title:
            hard.append("Chybi nazev produktu (Title)")
            hard_fields.add("title")

        if row.get("price") is None or (isinstance(row.get("price"), float) and pd.isna(row.get("price"))):
            hard.append("Chybi cena (Price)")
            hard_fields.add("price")
        elif price_val is None:
            hard.append(f"Cenu se nepodarilo rozpoznat jako cislo: {row.get('price')!r}")
            hard_fields.add("price")
        elif price_val <= 0:
            hard.append(f"Cena musi byt kladne cislo (nalezeno: {price_val})")
            hard_fields.add("price")
        elif isinstance(row.get("price"), str):
            soft.append(f"Cena zadana v nestandardnim formatu '{row.get('price')}' - automaticky rozpoznana hodnota {price_val}, prosim overit")
            soft_fields.add("price")

        if not sku:
            hard.append("Chybi SKU")
            hard_fields.add("sku")
        elif sku_counts.get(sku, 0) > 1:
            hard.append(f"Duplicitni SKU '{sku}' (vyskyt {sku_counts[sku]}x ve vstupu)")
            hard_fields.add("sku")
        elif idx in auto_deduped:
            soft.append(f"SKU automaticky upraveno na '{sku}' kvuli duplicite s jinym radkem (puvodne '{auto_deduped[idx]}') - prosim potvrdit spravnost")
            soft_fields.add("sku")

        if title:
            h = slugify(title)
            if handle_counts.get(h, 0) > 1 and not (sku and sku_counts.get(sku, 0) > 1):
                hard.append(f"Duplicitni nazev produktu / handle '{h}' pouzity vicekrat")
                hard_fields.add("title")

        if not vendor:
            soft.append(f"Chybi znacka (Vendor) - doplneno vychozi '{STORE_DEFAULT_VENDOR}', prosim potvrdit")
            soft_fields.add("vendor")
        if not category:
            soft.append("Chybi kategorie - AI navrhla kategorii na zaklade nazvu, prosim potvrdit")
            soft_fields.add("category")
        if not description:
            soft.append("Chybi popis produktu - vygenerovan AI navrh, prosim zkontrolovat")
            soft_fields.add("description")
        if not image_url:
            soft.append("Chybi URL obrazku - nutne doplnit rucne pred publikaci")
            soft_fields.add("image_url")

        if inv_qty is not None and not (isinstance(inv_qty, float) and pd.isna(inv_qty)):
            try:
                if float(inv_qty) <= 0:
                    soft.append("Skladem 0 ks - produkt bude importovan jako vyprodany")
                    soft_fields.add("inventory_qty")
            except (ValueError, TypeError):
                pass

        results.append({
            "index": idx, "hard_issues": hard, "soft_issues": soft, "price_clean": price_val,
            "hard_fields": hard_fields, "soft_fields": soft_fields,
        })
    return results


def flag_price_anomalies(df: pd.DataFrame, validations: list):
    by_category = {}
    for v in validations:
        if v["hard_issues"]:
            continue
        row = df.loc[v["index"]]
        cat = clean_text(row.get("category")) or "Uncategorized"
        if v["price_clean"]:
            by_category.setdefault(cat, []).append(v["price_clean"])

    medians = {cat: median(vals) for cat, vals in by_category.items() if len(vals) >= 2}

    for v in validations:
        row = df.loc[v["index"]]
        cat = clean_text(row.get("category")) or "Uncategorized"
        m = medians.get(cat)
        if m and v["price_clean"]:
            ratio = v["price_clean"] / m
            if ratio > 2.0 or ratio < 0.4:
                v["soft_issues"].append(
                    f"Cena {v['price_clean']} vybocuje z medianu kategorie '{cat}' ({m:.0f}) - prosim overit"
                )
                v["soft_fields"].add("price")
    return validations


def flag_sku_anomalies(df: pd.DataFrame, validations: list):
    """Odhali SKU, ktere tvarem vybocuji ze zbytku vstupu (napr. '123' mezi
    'FAB-STR-002', 'FAB-STR-003', ...) - typicky priznak preklepu nebo
    placeholderu, ktery samotna kontrola na prazdnotu/duplicitu neodhali."""
    skus = [clean_text(df.loc[v["index"]].get("sku")) for v in validations]
    skus = [s for s in skus if s]
    if len(skus) < 3:
        return validations

    has_letters_count = sum(1 for s in skus if re.search(r"[A-Za-z]", s))
    letters_are_typical = has_letters_count / len(skus) >= 0.7
    median_len = median(len(s) for s in skus)

    for v in validations:
        sku = clean_text(df.loc[v["index"]].get("sku"))
        if not sku:
            continue
        if letters_are_typical and not re.search(r"[A-Za-z]", sku):
            v["soft_issues"].append(
                f"SKU '{sku}' neodpovida formatu ostatnich SKU ve vstupu (chybi pismena) - prosim zkontrolovat"
            )
            v["soft_fields"].add("sku")
        elif median_len > 0 and len(sku) < median_len * 0.5:
            v["soft_issues"].append(
                f"SKU '{sku}' je nezvykle kratke oproti ostatnim SKU (median delky {int(median_len)} znaku) - prosim zkontrolovat"
            )
            v["soft_fields"].add("sku")
    return validations


def recompute_validations(df: pd.DataFrame) -> list:
    """Spusti cely retezec kontrol (zakladni validace + cenove a SKU anomalie)
    v jednom kroku - pouzito vsude tam, kde je potreba data po zmene znovu
    prevalidovat, aby retezec zustal na jednom miste."""
    return flag_sku_anomalies(df, flag_price_anomalies(df, validate_rows(df)))


def auto_fix_fixable(df: pd.DataFrame, validations: list, enriched: list) -> set:
    """Automaticky doplni to, co lze opravit bez rizika vymysleni spatnych dat:
    chybejici popis produktu se dopise do vstupnich dat AI vygenerovanym textem
    (uz se stejne pouziva v exportu), aby varovani "chybi popis" prestalo viset
    navzdy jen proto, ze puvodni vstup ho neobsahoval. SKU se NEvymysli - chybejici
    nebo duplicitni SKU vyzaduje vzdy rucni zasah, protoze falesne SKU muze zpusobit
    skutecny problem pri importu do Shopify.
    Vraci mnozinu indexu radku, ktere byly upraveny."""
    modified = set()

    for v, ai in zip(validations, enriched):
        idx = v["index"]
        row = df.loc[idx]

        if "description" in v["soft_fields"] and not clean_text(row.get("description")):
            plain = re.sub(r"<[^<]+?>", "", ai.get("ai_body_html", "") or "").strip()
            if plain:
                df.at[idx, "description"] = plain
                modified.add(idx)

    return modified


def flag_missing_ai_content(validations: list, enriched: list):
    """Bezpecnostni pojistka nad samotnym AI enrichmentem: ai_enrich_all normalne
    vygeneruje SEO title/description/tagy/popis pro kazdy produkt s nazvem, ale
    pokud presto neco chybi (typicky protoze radek nema nazev, na kterem AI obsah
    zavisi, nebo enrichment selhal), musi to byt videt jako varovani u konkretnich
    sloupcu - ne tise prazdne bunky bez vysvetleni."""
    for v, ai in zip(validations, enriched):
        missing_labels = []
        for field, label in (
            ("ai_seo_title", "AI SEO Title"),
            ("ai_seo_description", "AI SEO Description"),
            ("ai_tags", "AI Tags"),
            ("ai_body_html", "AI popis (HTML)"),
        ):
            if not clean_text(ai.get(field)):
                missing_labels.append(label)
                v["soft_fields"].add(field)
        if missing_labels:
            if ai.get("ai_mode") == "error":
                reason = f"AI enrichment selhal ({'; '.join(ai.get('ai_notes', [])) or 'neznama chyba'}) - zkuste tlacitko Automaticky opravit"
            else:
                reason = "obvykle kvuli chybejicimu nazvu produktu, zkontrolujte vstup"
            v["soft_issues"].append(f"Chybi AI obsah ({', '.join(missing_labels)}) - {reason}")
    return validations


# ---------------------------------------------------------------------------
# AI enrichment (marketing / SEO / AEO)
# ---------------------------------------------------------------------------

def display_price(raw) -> str:
    """Textova reprezentace puvodni (nevycistene) ceny pro zobrazeni v tabulkach -
    sloupec smesici cisla a texty by jinak shazoval Arrow serializaci v UI."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return str(raw)


def suggest_category(title: str) -> str:
    t = title.lower()
    for kw, cat in CATEGORY_KEYWORDS:
        if kw in t:
            return cat
    return "General"


EMPTY_AI_RESULT = {
    "ai_body_html": "", "ai_seo_title": "", "ai_seo_description": "",
    "ai_seo_keywords": "", "ai_tags": "", "ai_suggested_category": "",
    "ai_faq": [], "ai_aeo_answer": "", "ai_notes": [], "ai_mode": "mock",
}


def _mock_one(p: dict) -> dict:
    title, vendor, category = p["title"], p["vendor"], p["category"]
    description, color, material = p["description"], p["color"], p["material"]

    cat = category or suggest_category(title)
    base_desc = description
    if not base_desc:
        parts = [x for x in [material, color] if x]
        detail = f" v provedeni {', '.join(parts)}" if parts else ""
        base_desc = f"{title}{detail}. Kvalitni zpracovani a pohodlny kazdodenni nosnost z kolekce {vendor or STORE_DEFAULT_VENDOR}."

    aeo_answer = f"{title} od {vendor or STORE_DEFAULT_VENDOR} je produkt z kategorie {cat}, vhodny pro kazdodenni pouziti" + (f", vyrobeny z {material}." if material else ".")

    body_html = (
        f"<p><strong>{aeo_answer}</strong></p>"
        f"<p>{base_desc}</p>"
        "<ul>" + "".join(f"<li>{label}: {value}</li>" for label, value in
                          [("Kategorie", cat), ("Material", material), ("Barva", color)] if value) + "</ul>"
    )

    seo_title = f"{title} | {STORE_NAME_FOR_SEO}"
    if len(seo_title) > 70:
        seo_title = seo_title[:67].rstrip() + "..."

    seo_description = re.sub("<[^<]+?>", "", base_desc)
    seo_description = (seo_description[:157] + "...") if len(seo_description) > 160 else seo_description

    tags = sorted({t for t in [cat, color, material, "new-collection"] if t})
    keywords = sorted({t.lower() for t in [title, cat, color, material, vendor] if t})

    faq = [
        {"question": f"Z jakeho materialu je {title.lower()}?",
         "answer": f"{title} je vyrobeny z {material}." if material else f"Material u produktu {title} je uveden na produktove strance."},
        {"question": f"Jaka je cena produktu {title}?",
         "answer": "Aktualni cenu najdete primo na produktove strance."},
        {"question": f"Pro koho je {title.lower()} vhodny?",
         "answer": aeo_answer},
    ]

    return {
        "ai_body_html": body_html,
        "ai_seo_title": seo_title,
        "ai_seo_description": seo_description,
        "ai_seo_keywords": ", ".join(keywords),
        "ai_tags": ", ".join(tags),
        "ai_suggested_category": cat if not category else category,
        "ai_faq": faq,
        "ai_aeo_answer": aeo_answer,
        "ai_notes": [],
        "ai_mode": "mock",
    }


def _call_anthropic_batch(products: list) -> str:
    import urllib.request
    api_key = os.environ["ANTHROPIC_API_KEY"]

    items = "\n".join(
        f"{i}. Nazev: {p['title']} | Kategorie: {p['category'] or '(chybi)'} | "
        f"Popis: {p['description'] or '(chybi)'} | Barva: {p['color']} | Material: {p['material']} | "
        f"Cena: {p['price']}"
        for i, p in enumerate(products)
    )

    prompt = f"""Jsi e-commerce copywriter a SEO/AEO specialista pro znacku "{STORE_NAME_FOR_SEO}"
(AEO = Answer Engine Optimization - obsah ma byt snadno citovatelny AI asistenty
jako ChatGPT, Perplexity nebo Google AI Overviews, tedy strukturovany a s jasnymi
primymi odpovedmi).

Pro kazdy z nasledujicich produktu vygeneruj obsah pro Shopify import. Vrat VYHRADNE
validni JSON pole (bez dalsiho textu) se stejnym poctem prvku ve stejnem poradi, kde
kazdy prvek ma klice:
  body_html (2-4 vety popisu jako HTML, prvni veta jako kratka "primo odpoved" na
    otazku "co to je a pro koho to je" - obalena v <strong>),
  seo_title (max 70 znaku),
  seo_description (max 160 znaku),
  seo_keywords (pole 5-8 klicovych slov/fraz pro SEO),
  tags (pole 3-6 tagu),
  suggested_category (pokud kategorie chybi, navrhni; jinak vrat puvodni),
  faq (pole 2-3 objektu {{question, answer}} - bezne zakaznicke otazky k produktu,
    kratke faktualni odpovedi vhodne pro AEO/featured snippets),
  aeo_answer (1 veta, primy vecny popis produktu vhodny jako citace v AI odpovedi),
  notes (pole retezcu - cokoliv, co ti prijde u produktu podezrele nebo nekonzistentni,
    napr. nesedici nazev/kategorie, prilis vysoka/nizka cena, prazdny prvek pokud nic).

Produkty:
{items}
"""

    body = json.dumps({
        "model": ANTHROPIC_MODEL,
        "max_tokens": 4000,
        "messages": [{"role": "user", "content": prompt}],
    }).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=body,
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read())
    return data["content"][0]["text"]


def _parse_ai_batch_response(raw: str, expected_len: int):
    match = re.search(r"\[.*\]", raw, re.DOTALL)
    if not match:
        raise ValueError("Odpoved AI neobsahuje JSON pole")
    items = json.loads(match.group(0))
    if not isinstance(items, list) or len(items) != expected_len:
        raise ValueError(f"Neocekavany pocet polozek v AI odpovedi ({len(items) if isinstance(items, list) else 'N/A'} misto {expected_len})")
    return items


def ai_enrich_all(df: pd.DataFrame, validations: list, progress_cb=None) -> list:
    """Vrati list AI vysledku (stejna delka a poradi jako `validations`)."""
    mode = get_ai_mode()
    products = []
    for v in validations:
        row = df.loc[v["index"]]
        products.append({
            "title": clean_text(row.get("title")),
            "vendor": clean_text(row.get("vendor")),
            "category": clean_text(row.get("category")),
            "description": clean_text(row.get("description")),
            "color": clean_text(row.get("color")),
            "material": clean_text(row.get("material")),
            "price": v["price_clean"],
        })

    results = [None] * len(products)

    # produkty bez nazvu AI nezpracovava
    to_process_idx = [i for i, p in enumerate(products) if p["title"]]
    for i in range(len(products)):
        if i not in to_process_idx:
            results[i] = dict(EMPTY_AI_RESULT, ai_mode=mode)

    if mode == "mock":
        for i in to_process_idx:
            results[i] = _mock_one(products[i])
            if progress_cb:
                progress_cb(i + 1, len(products))
        return results

    # realny AI rezim - davkove zpracovani
    batches = [to_process_idx[i:i + AI_BATCH_SIZE] for i in range(0, len(to_process_idx), AI_BATCH_SIZE)]
    done = 0
    for batch in batches:
        batch_products = [products[i] for i in batch]
        try:
            raw = _call_anthropic_batch(batch_products)
            parsed = _parse_ai_batch_response(raw, len(batch_products))
            for i, item in zip(batch, parsed):
                results[i] = {
                    "ai_body_html": item.get("body_html", ""),
                    "ai_seo_title": item.get("seo_title", ""),
                    "ai_seo_description": item.get("seo_description", ""),
                    "ai_seo_keywords": ", ".join(item.get("seo_keywords", []) or []),
                    "ai_tags": ", ".join(item.get("tags", []) or []),
                    "ai_suggested_category": item.get("suggested_category") or products[i]["category"],
                    "ai_faq": item.get("faq", []) or [],
                    "ai_aeo_answer": item.get("aeo_answer", ""),
                    "ai_notes": item.get("notes", []) or [],
                    "ai_mode": "anthropic",
                }
        except Exception as exc:
            # Zadny tichy fallback na mock text - selhani davky musi byt videt.
            # Radek zustane bez AI obsahu, coz `flag_missing_ai_content` odhali
            # a oznaci jako varovani; tlacitko "Automaticky opravit" pak muze
            # enrichment pro tyto radky zopakovat.
            for i in batch:
                results[i] = dict(EMPTY_AI_RESULT, ai_mode="error", ai_notes=[f"AI enrichment selhal: {exc}"])
        done += len(batch)
        if progress_cb:
            progress_cb(done, len(to_process_idx))

    return results


def revalidate_and_reenrich(df: pd.DataFrame, old_enriched: list, edited_indices, progress_cb=None):
    """Pouziva se po rucni oprave radku ve webove appce: prevaliduje CELY soubor
    (kvuli globalnim kontrolam jako duplicitni SKU/handle) a AI enrichment znovu
    spusti jen pro radky v `edited_indices` - ostatni si ponechaji puvodni AI vysledek
    (setri to cas i naklady na API). Vraci (validations, enriched) se stejnym poradim
    jako `validate_rows(df)`.
    """
    new_validations = recompute_validations(df)
    new_enriched = list(old_enriched)
    edited_indices = set(edited_indices)

    to_reenrich_positions = [i for i, v in enumerate(new_validations) if v["index"] in edited_indices]
    if to_reenrich_positions:
        subset_validations = [new_validations[i] for i in to_reenrich_positions]
        subset_enriched = ai_enrich_all(df, subset_validations, progress_cb=progress_cb)
        for i, ai in zip(to_reenrich_positions, subset_enriched):
            new_enriched[i] = ai

    new_validations = flag_missing_ai_content(new_validations, new_enriched)
    return new_validations, new_enriched


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

# Ktery interni nazev pole odpovida ktere viditelne QC-sloupci - pouzito pro
# cilene zvyrazneni presne te bunky, kde je problem, misto celeho radku.
FIELD_TO_QC_COLUMN = {
    "title": "Title",
    "vendor": "Vendor",
    "category": "Category (input)",
    "price": "Price (input)",
    "sku": "SKU",
    "image_url": "Image URL",
    "description": "Description (input)",
    "inventory_qty": "Stock (input)",
    "ai_seo_title": "AI SEO Title",
    "ai_seo_description": "AI SEO Description",
    "ai_tags": "AI Tags",
    "ai_body_html": "AI Generated Description (HTML)",
}


def build_qc_dataframe(df: pd.DataFrame, validations: list, enriched: list) -> pd.DataFrame:
    rows = []
    for v, ai in zip(validations, enriched):
        row = df.loc[v["index"]]
        ready = len(v["hard_issues"]) == 0
        all_soft = list(v["soft_issues"]) + [f"AI: {n}" for n in ai.get("ai_notes", [])]
        rows.append({
            "Ready_for_Import": "ANO" if ready else "NE",
            "Title": row.get("title"),
            "Vendor": clean_text(row.get("vendor")) or (STORE_DEFAULT_VENDOR if ready else ""),
            "Category (input)": row.get("category"),
            "AI Suggested Category": ai["ai_suggested_category"],
            "Price (input)": display_price(row.get("price")),
            "SKU": row.get("sku"),
            "Stock (input)": row.get("inventory_qty"),
            "Image URL": row.get("image_url"),
            "Description (input)": row.get("description"),
            "Blocking Issues": " | ".join(v["hard_issues"]) if v["hard_issues"] else "",
            "Warnings / Review": " | ".join(all_soft) if all_soft else "",
            "AI SEO Title": ai["ai_seo_title"],
            "AI SEO Description": ai["ai_seo_description"],
            "AI SEO Keywords": ai["ai_seo_keywords"],
            "AI Tags": ai["ai_tags"],
            "AEO Answer Snippet": ai["ai_aeo_answer"],
            "AI FAQ": json.dumps(ai["ai_faq"], ensure_ascii=False),
            "AI Generated Description (HTML)": ai["ai_body_html"],
            "AI Mode": ai["ai_mode"],
            "_hard_fields": ",".join(sorted(v["hard_fields"])),
            "_soft_fields": ",".join(sorted(v["soft_fields"])),
        })
    return pd.DataFrame(rows)


def qc_dataframe_to_xlsx_bytes(qc_df: pd.DataFrame) -> bytes:
    """Serializuje QC report do stylovaneho .xlsx (v pameti, bez zapisu na disk).
    Misto barveni cele radky zvyrazni jen konkretni bunku(ky), kde je problem -
    cervene pro blokujici chybu, zlute pro pouhe varovani - aby bylo na prvni
    pohled videt CO presne je spatne, ne jen ktery radek."""
    hard_fields_col = qc_df["_hard_fields"] if "_hard_fields" in qc_df.columns else pd.Series([""] * len(qc_df))
    soft_fields_col = qc_df["_soft_fields"] if "_soft_fields" in qc_df.columns else pd.Series([""] * len(qc_df))
    visible_df = qc_df.drop(columns=[c for c in ["_hard_fields", "_soft_fields"] if c in qc_df.columns])

    buf = io.BytesIO()
    visible_df.to_excel(buf, index=False, sheet_name="QC Report", engine="openpyxl")
    buf.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(buf)
    ws = wb["QC Report"]
    hard_fill = PatternFill(start_color="FFF2CBCB", end_color="FFF2CBCB", fill_type="solid")
    soft_fill = PatternFill(start_color="FFFCE8B2", end_color="FFFCE8B2", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    col_letter_by_name = {cell.value: cell.column_letter for cell in ws[1]}

    for i, row_idx in enumerate(range(2, ws.max_row + 1)):
        hard_field_names = [f for f in hard_fields_col.iloc[i].split(",") if f]
        soft_field_names = [f for f in soft_fields_col.iloc[i].split(",") if f]
        for field in hard_field_names:
            col_name = FIELD_TO_QC_COLUMN.get(field)
            letter = col_letter_by_name.get(col_name)
            if letter:
                ws[f"{letter}{row_idx}"].fill = hard_fill
        for field in soft_field_names:
            if field in hard_field_names:
                continue  # blokujici barva ma prednost pred varovnou na stejne bunce
            col_name = FIELD_TO_QC_COLUMN.get(field)
            letter = col_letter_by_name.get(col_name)
            if letter:
                ws[f"{letter}{row_idx}"].fill = soft_fill

    for col_idx, col_name in enumerate(visible_df.columns, start=1):
        width = min(60, max(12, visible_df[col_name].astype(str).str.len().quantile(0.9) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def shopify_dataframe_to_csv_bytes(shopify_df: pd.DataFrame) -> bytes:
    return shopify_df.to_csv(index=False).encode("utf-8-sig")


def build_shopify_dataframe(df: pd.DataFrame, validations: list, enriched: list, include_blocked: bool = False) -> pd.DataFrame:
    """include_blocked=False (vychozi): jen radky pripravene k importu (bez
    blokujicich chyb) - to co se skutecne ma poslat do Shopify. include_blocked=True:
    vsechny radky bez ohledu na stav - pro kontrolu, co presne by import obsahoval."""
    shopify_rows = []
    for v, ai in zip(validations, enriched):
        if v["hard_issues"] and not include_blocked:
            continue
        row = df.loc[v["index"]]
        title = clean_text(row.get("title"))
        handle = slugify(title)
        vendor = clean_text(row.get("vendor")) or STORE_DEFAULT_VENDOR
        category = ai["ai_suggested_category"] or clean_text(row.get("category"))
        image_url = clean_text(row.get("image_url"))
        inv_qty = row.get("inventory_qty")
        try:
            inv_qty = int(inv_qty) if inv_qty is not None and not pd.isna(inv_qty) else 0
        except (ValueError, TypeError):
            inv_qty = 0
        weight = row.get("weight_g")
        try:
            weight = float(weight) if weight is not None and not pd.isna(weight) else ""
        except (ValueError, TypeError):
            weight = ""

        shopify_rows.append({
            "Handle": handle,
            "Title": title,
            "Body (HTML)": ai["ai_body_html"],
            "Vendor": vendor,
            "Product Category": category,
            "Type": category,
            "Tags": ai["ai_tags"],
            "Published": "TRUE",
            "Option1 Name": "Title",
            "Option1 Value": "Default Title",
            "Variant SKU": clean_text(row.get("sku")),
            "Variant Grams": weight,
            "Variant Inventory Tracker": "shopify",
            "Variant Inventory Qty": inv_qty,
            "Variant Inventory Policy": "deny",
            "Variant Fulfillment Service": "manual",
            "Variant Price": v["price_clean"],
            "Variant Requires Shipping": "TRUE",
            "Variant Taxable": "TRUE",
            "Image Src": image_url,
            "Image Position": 1 if image_url else "",
            "Image Alt Text": title,
            "SEO Title": ai["ai_seo_title"],
            "SEO Description": ai["ai_seo_description"],
            "Status": "active",
            "Metafield: custom.seo_keywords [list.single_line_text_field]": ai["ai_seo_keywords"],
            "Metafield: custom.faq [json]": json.dumps(ai["ai_faq"], ensure_ascii=False),
            "Metafield: custom.aeo_answer [single_line_text_field]": ai["ai_aeo_answer"],
        })

    return pd.DataFrame(shopify_rows)


def run_pipeline(input_path_or_buffer, progress_cb=None):
    """Spusti cely pipeline a vrati (df, validations, enriched, qc_df, shopify_df)."""
    df = load_input(input_path_or_buffer)
    validations = recompute_validations(df)
    enriched = ai_enrich_all(df, validations, progress_cb=progress_cb)
    validations = flag_missing_ai_content(validations, enriched)
    qc_df = build_qc_dataframe(df, validations, enriched)
    shopify_df = build_shopify_dataframe(df, validations, enriched)
    return df, validations, enriched, qc_df, shopify_df
